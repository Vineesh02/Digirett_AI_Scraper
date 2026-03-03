"""
FIX EXISTING DATABASE RECORDS
This script will:
1. Export all existing data to Excel with content preview
2. Show statistics about current data
3. Optionally update category/subcategory for files in wrong categories
"""

import os
import sys
import logging
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

# Add parent directory to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from database import Database
from storage_handler import StorageHandler
import config

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)


class DatabaseFixer:
    """
    Fix and export existing database records
    """
    
    def __init__(self):
        self.db = Database()
        self.storage = StorageHandler()
    
    # ======================================================
    # EXPORT TO EXCEL
    # ======================================================
    
    def export_complete_excel(self, output_file: str = None):
        """
        Export ALL database records to Excel with content preview
        """
        if not output_file:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_file = f"lovdata_complete_export_{timestamp}.xlsx"
        
        logger.info("=" * 80)
        logger.info("EXPORTING COMPLETE DATABASE TO EXCEL")
        logger.info("=" * 80)
        
        # Get all metadata
        logger.info("📥 Fetching all records from database...")
        records = self.db.get_all_metadata()
        
        if not records:
            logger.error("❌ No records found in database!")
            return
        
        logger.info(f"✓ Found {len(records)} records")
        
        # Prepare data for Excel
        excel_data = []
        
        for i, record in enumerate(records, 1):
            if i % 50 == 0:
                logger.info(f"  Processing record {i}/{len(records)}...")
            
            # Get content preview from XML file or storage
            content_preview = self._get_content_preview(record)
            
            excel_data.append({
                "Filename": record.get("file_name", ""),
                "URL": record.get("source_url", ""),
                "Category": record.get("category", ""),
                "Subcategory": record.get("subcategory", ""),
                "Content Preview (500 chars)": content_preview[:500] if content_preview else "",
                "File Size (KB)": round(record.get("file_size", 0) / 1024, 2),
                "Public URL": record.get("public_url", ""),
                "Storage Path": record.get("storage_path", ""),
                "Created At": record.get("created_at", ""),
            })
        
        # Create DataFrame
        df = pd.DataFrame(excel_data)
        
        # Export to Excel
        logger.info(f"📊 Writing to Excel: {output_file}")
        df.to_excel(output_file, index=False, engine='openpyxl')
        
        # Apply formatting
        self._format_excel(output_file)
        
        logger.info(f"✅ Excel export complete: {output_file}")
        logger.info(f"   Total records: {len(excel_data)}")
        
        return output_file
    
    # ======================================================
    # GET CONTENT PREVIEW
    # ======================================================
    
    def _get_content_preview(self, record: dict) -> str:
        """
        Get content preview from XML file
        Try local file first, then download from Supabase if needed
        """
        try:
            # Try to get from metadata first (if it exists)
            if record.get("content_preview"):
                return record["content_preview"]
            
            # Try local file
            file_name = record.get("file_name")
            category = record.get("category", "Unknown")
            subcategory = record.get("subcategory", "Unknown")
            
            local_path = os.path.join(
                config.BASE_DIR,
                category,
                subcategory,
                file_name
            )
            
            if os.path.exists(local_path):
                with open(local_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                    # Extract text content (simple approach)
                    from bs4 import BeautifulSoup
                    soup = BeautifulSoup(content, 'xml')
                    text = soup.get_text()
                    return ' '.join(text.split())  # Normalize whitespace
            
            # If no local file, try to download from Supabase
            # (This would require implementing download logic)
            
            return "Content not available"
            
        except Exception as e:
            logger.debug(f"Could not get content preview: {e}")
            return "Content preview error"
    
    # ======================================================
    # FORMAT EXCEL
    # ======================================================
    
    def _format_excel(self, filepath: str):
        """
        Apply formatting to Excel file
        """
        try:
            wb = load_workbook(filepath)
            ws = wb.active
            ws.title = "Lovdata Documents"
            
            # Header styling
            header_fill = PatternFill("solid", fgColor="2F5496")
            header_font = Font(bold=True, color="FFFFFF")
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
            
            # Apply to header row
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            
            # Freeze panes
            ws.freeze_panes = "C2"
            
            # Column widths
            ws.column_dimensions["A"].width = 35  # Filename
            ws.column_dimensions["B"].width = 60  # URL
            ws.column_dimensions["C"].width = 30  # Category
            ws.column_dimensions["D"].width = 30  # Subcategory
            ws.column_dimensions["E"].width = 80  # Content Preview
            ws.column_dimensions["F"].width = 15  # File Size
            ws.column_dimensions["G"].width = 60  # Public URL
            ws.column_dimensions["H"].width = 50  # Storage Path
            ws.column_dimensions["I"].width = 22  # Created At
            
            wb.save(filepath)
            logger.info("✓ Excel formatting applied")
            
        except Exception as e:
            logger.warning(f"⚠ Could not apply formatting: {e}")
    
    # ======================================================
    # STATISTICS
    # ======================================================
    
    def show_statistics(self):
        """
        Show database statistics
        """
        logger.info("=" * 80)
        logger.info("DATABASE STATISTICS")
        logger.info("=" * 80)
        
        stats = self.db.get_statistics()
        
        logger.info(f"\nTotal Files: {stats['total_files']}")
        logger.info(f"Total Size: {stats['total_size_mb']:.2f} MB")
        
        logger.info("\nBy Category:")
        for cat, data in sorted(stats['by_category'].items()):
            logger.info(f"  {cat}: {data['count']} files ({data['size_mb']:.2f} MB)")
        
        logger.info("\nBy Subcategory:")
        for subcat, data in sorted(stats['by_subcategory'].items()):
            logger.info(f"  {subcat}: {data['count']} files ({data['size_mb']:.2f} MB)")
        
        logger.info("=" * 80)
    
    # ======================================================
    # FIND DUPLICATE FILENAMES
    # ======================================================
    
    def find_duplicates(self):
        """
        Find files that might be duplicated due to category mismatches
        """
        logger.info("=" * 80)
        logger.info("CHECKING FOR POTENTIAL DUPLICATES")
        logger.info("=" * 80)
        
        all_records = self.db.get_all_metadata()
        
        # Group by filename
        filename_groups = {}
        for record in all_records:
            fname = record.get("file_name")
            if fname:
                if fname not in filename_groups:
                    filename_groups[fname] = []
                filename_groups[fname].append(record)
        
        # Find duplicates
        duplicates = {k: v for k, v in filename_groups.items() if len(v) > 1}
        
        if duplicates:
            logger.info(f"\n⚠ Found {len(duplicates)} duplicate filenames:")
            for fname, records in duplicates.items():
                logger.info(f"\n  {fname}:")
                for rec in records:
                    logger.info(f"    - Category: {rec.get('category')}, Subcategory: {rec.get('subcategory')}")
        else:
            logger.info("✓ No duplicate filenames found")
        
        logger.info("=" * 80)


# ============================================================
# MAIN
# ============================================================

def main():
    """
    Main function with menu
    """
    fixer = DatabaseFixer()
    
    print("\n" + "=" * 80)
    print("LOVDATA DATABASE FIXER & EXPORTER")
    print("=" * 80)
    print("\nOptions:")
    print("1. Export complete database to Excel")
    print("2. Show database statistics")
    print("3. Find duplicate filenames")
    print("4. Do all of the above")
    print("5. Exit")
    print()
    
    choice = input("Select option (1-5): ").strip()
    
    if choice == "1":
        output_file = input("\nEnter output filename (or press Enter for default): ").strip()
        if not output_file:
            output_file = None
        fixer.export_complete_excel(output_file)
    
    elif choice == "2":
        fixer.show_statistics()
    
    elif choice == "3":
        fixer.find_duplicates()
    
    elif choice == "4":
        fixer.show_statistics()
        print()
        fixer.find_duplicates()
        print()
        output_file = input("\nEnter output filename for Excel export (or press Enter for default): ").strip()
        if not output_file:
            output_file = None
        fixer.export_complete_excel(output_file)
    
    elif choice == "5":
        print("\nGoodbye!")
        return
    
    else:
        print("\nInvalid option!")
        return
    
    print("\n" + "=" * 80)
    print("DONE!")
    print("=" * 80)


if __name__ == "__main__":
    main()