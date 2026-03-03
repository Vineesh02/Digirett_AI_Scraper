# """
# Excel Exporter
# Reads CONTENT from XML (not DB)
# """

# import os
# import pandas as pd
# import logging
# from typing import List, Dict
# from datetime import datetime
# from openpyxl import load_workbook
# from openpyxl.styles import Font, PatternFill, Alignment
# from xml_handler import XMLHandler
# import config

# logger = logging.getLogger(__name__)


# class ExcelExporter:

#     HEADERS = [
#         "Filename",
#         "Category",
#         "Subcategory",
#         "Source URL",
#         "File Size (KB)",
#         "File Hash",
#         "Content Preview"
#     ]

#     def export(self, metadata_rows: List[Dict]) -> str:
#         rows = []

#         for meta in metadata_rows:
#             xml_path = meta.get("storage_path")

#             preview = (
#                 XMLHandler.read_content_preview(xml_path)
#                 if xml_path and os.path.exists(xml_path)
#                 else ""
#             )

#             rows.append({
#                 "Filename": meta.get("file_name"),
#                 "Category": meta.get("category"),
#                 "Subcategory": meta.get("subcategory"),
#                 "Source URL": meta.get("source_url"),
#                 "File Size (KB)": round((meta.get("file_size") or 0) / 1024, 2),
#                 "File Hash": meta.get("file_hash"),
#                 "Content Preview": preview
#             })

#         df = pd.DataFrame(rows)

#         os.makedirs(config.EXCEL_DIR, exist_ok=True)
#         file_path = os.path.join(
#             config.EXCEL_DIR,
#             f"lovdata_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
#         )

#         df.to_excel(file_path, index=False, engine="openpyxl")
#         self._format(file_path)

#         logger.info(f"✓ Excel created: {file_path}")
#         return file_path

#     def _format(self, path: str):
#         wb = load_workbook(path)
#         ws = wb.active

#         header_fill = PatternFill("solid", fgColor="2F5496")
#         header_font = Font(bold=True, color="FFFFFF")

#         for col in range(1, ws.max_column + 1):
#             cell = ws.cell(row=1, column=col)
#             cell.fill = header_fill
#             cell.font = header_font
#             ws.column_dimensions[cell.column_letter].width = 35
#             cell.alignment = Alignment(wrap_text=True)

#         ws.freeze_panes = "A2"
#         ws.auto_filter.ref = ws.dimensions
#         wb.save(path)
"""
Excel Exporter - READS CONTENT FROM SUPABASE
Downloads XML from public_url and extracts content preview
"""

import os
import pandas as pd
import logging
import requests
import xml.etree.ElementTree as ET
from typing import List, Dict
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import config

logger = logging.getLogger(__name__)


class ExcelExporter:

    HEADERS = [
        "Filename",
        "Category",
        "Subcategory",
        "Source URL",
        "File Size (KB)",
        "File Hash",
        "Content Preview"
    ]

    def export(self, metadata_rows: List[Dict]) -> str:
        """Export metadata to Excel with content preview from Supabase"""
        rows = []

        logger.info(f"📊 Processing {len(metadata_rows)} documents for Excel...")

        for i, meta in enumerate(metadata_rows, 1):
            if i % 50 == 0:
                logger.info(f"  Progress: {i}/{len(metadata_rows)}")
            
            # Get content preview from Supabase public URL
            preview = self._get_content_preview_from_url(meta.get("public_url"))

            rows.append({
                "Filename": meta.get("file_name"),
                "Category": meta.get("category"),
                "Subcategory": meta.get("subcategory"),
                "Source URL": meta.get("source_url"),
                "File Size (KB)": round((meta.get("file_size") or 0) / 1024, 2),
                "File Hash": meta.get("file_hash"),
                "Content Preview": preview
            })

        df = pd.DataFrame(rows)

        os.makedirs(config.EXCEL_DIR, exist_ok=True)
        file_path = os.path.join(
            config.EXCEL_DIR,
            f"lovdata_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        df.to_excel(file_path, index=False, engine="openpyxl")
        self._format(file_path)

        logger.info(f"✓ Excel created: {file_path}")
        return file_path

    def _get_content_preview_from_url(self, public_url: str, max_chars: int = 500) -> str:
        """
        Download XML from Supabase public URL and extract content preview.
        
        Args:
            public_url: Supabase public URL to XML file
            max_chars: Maximum characters for preview
            
        Returns:
            First 500 chars of content, or empty string if error
        """
        if not public_url:
            return ""
        
        try:
            # Download XML from Supabase
            response = requests.get(public_url, timeout=10)
            response.raise_for_status()
            
            # Parse XML
            root = ET.fromstring(response.content)
            
            # Find content element
            content_elem = root.find("content")
            if content_elem is not None and content_elem.text:
                full_content = content_elem.text.strip()
                return full_content[:max_chars]
            
            return ""
            
        except requests.exceptions.RequestException as e:
            logger.warning(f"  ⚠️ Failed to download {public_url}: {e}")
            return ""
        except ET.ParseError as e:
            logger.warning(f"  ⚠️ Failed to parse XML from {public_url}: {e}")
            return ""
        except Exception as e:
            logger.warning(f"  ⚠️ Error reading content from {public_url}: {e}")
            return ""

    def _format(self, path: str):
        """Format Excel file with headers and styling"""
        wb = load_workbook(path)
        ws = wb.active

        header_fill = PatternFill("solid", fgColor="2F5496")
        header_font = Font(bold=True, color="FFFFFF")

        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            ws.column_dimensions[cell.column_letter].width = 35
            cell.alignment = Alignment(wrap_text=True)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        wb.save(path)
    def append_row(self, meta: dict, content_preview: str = ""):
        """
        Append one row to Excel file
        """

        row = {
            "file_name": meta.get("file_name", ""),
            "category": meta.get("category", ""),
            "subcategory": meta.get("subcategory", ""),
            "source_url": meta.get("source_url", ""),
            "file_hash": meta.get("file_hash", ""),
            "file_size": meta.get("file_size", ""),
            "public_url": meta.get("public_url", ""),
            "storage_path": meta.get("bucket_path", ""),
            "content_preview": content_preview
        }

        # Now write row to Excel (your existing logic here)
